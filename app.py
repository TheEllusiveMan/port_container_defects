import os
import io
import csv
import time
import shutil
import streamlit as st
from PIL import Image as PILImage
from openpyxl import load_workbook

from models import *

st.set_page_config(page_title='HelpTallyman', page_icon=None, layout="centered", initial_sidebar_state="auto",
                   menu_items=None)
st.markdown("""
<style>
/* Изменение фона */
.stApp {
    background: #f5f5f5;
}
/* Стиль заголовков */
h1 {
    color: #0056b3;
    font-family: 'Arial';
}

#ee18af88 {
    color: #0056b3;
    font-family: 'Arial';
    font-size: calc(1rem + 1vw);
}

/* Header */
.stAppHeader {
    background: #0056b3;
}

.st-emotion-cache-uef7qa {
    color: #000000;
}

div.stElementContainer:nth-child(2) > div:nth-child(1) > div:nth-child(1) {
    color: #000000;
}    

/* Кнопка Найти поврждения */
button.st-emotion-cache-15hul6a:nth-child(1) {
    background-color: #0056b3;
}    

/* Коробка загрузки изображений */
.st-emotion-cache-1erivf3 {
    background-color: #0056b3;
}

/* Кнопка Browse files */
button.st-emotion-cache-15hul6a:nth-child(3) {
    background-color: #f5f5f5;
    color: #0056b3;
}

/* Список загруженных изображений */
.st-emotion-cache-fis6aj {
    background-color: #0056b3;
    color: #000000;
}

.st-emotion-cache-ltfnpr {
    background-color: #0056b3;
    color: #000000;
}
/* Кнопка Сформировать отчет */
div.stElementContainer:nth-child(7) > div:nth-child(1) {
    background-color: #0056b3;
}

# div.stElementContainer:nth-child(9) > div:nth-child(1)
# div.stElementContainer:nth-child(9) > div:nth-child(1) {
#     background-color: #0056b3;
# }

div.stElementContainer:nth-child(11) > div:nth-child(1) {
    background-color: #0056b3;
}

/* Sidebar */
.st-emotion-cache-6qob1r {
    background-color: #0056b3;
    # border-color: #0056b3;
    # border-width: 10px;
}

/* Надпись Выберите вариант в sidebar */
.stRadio > label:nth-child(1) > div:nth-child(1) {
    color: #f5f5f5;
}

/* Надпись Выберите изображение */
.stFileUploader > label:nth-child(1) > div:nth-child(1) {
    color: #0056b3;
}

/* Поле для ввода правильного номера контейнера */
#text_input_1 {
    color: #000000;
    background-color: #f5f5f5;
}

#text_input_2 {
    color: #000000;
    background-color: #f5f5f5;
}

#text_input_3 {
    color: #000000;
    background-color: #f5f5f5;
}

/* Надпись поля для ввода правильного номера контейнера */
.stTextInput > label:nth-child(1) > div:nth-child(1) {
    color: #0056b3;
}

</style>
""", unsafe_allow_html=True)

model_container = ModelContainer()
model_container_number = ModelContainerNumber()
model_container_damage = ModelContainerDamage()

# cont_result = []
dmg_translate = {'Deframe': 'деформация',
                 'Hole': 'пробоина',
                 'Rusty': 'ржавчина',
                 'Dent': 'вмятина',
                 'Scratch': 'царапины'}

wall_dict = {'Передняя стенка': 'на передней стенке ',
             'Задняя стенка': 'на задней стенке ',
             'Левая стенка': 'на левой стенке ',
             'Правая стенка': 'на правой стенке '}


def get_wall_str(list_result_list_number, side_str):
    if list_result_list_number == 0:
        side_str = wall_dict[st.session_state.wall_type_1_img]
    elif list_result_list_number == 1:
        side_str = wall_dict[st.session_state.wall_type_2_img]
    elif list_result_list_number == 2:
        side_str = wall_dict[st.session_state.wall_type_3_img]
    return side_str


def paint_results_dmg_number(cropped_img, results):
    # Отобразите результаты
    for result_list in results:
        for result in result_list:
            boxes = result.boxes  # Получите ограничивающие рамки
            names = result.names
            for box in boxes:
                x1, y1, x2, y2 = box.xyxy[0]  # Координаты bbox
                conf = box.conf[0]  # Уверенность
                cls = box.cls[0]  # Класс
                class_name = names[int(cls)]

                # Нарисуйте bbox на изображении
                cv2.rectangle(cropped_img, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 5)  # Красный цвет
                cv2.putText(cropped_img, f'Class: {class_name}, Conf: {conf:.2f}', (int(x1), int(y1) - 5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

    # st.write("Найденные повреждения для:", im_name)
    st.image(cropped_img, use_column_width=True)


def get_cropped_img_by_xyxys(image_rgb, xyxys):
    xyxy_1 = xyxys[0]
    xyxy = xyxy_1[0]
    x_min, y_min, x_max, y_max = xyxy[0], xyxy[1], xyxy[2], xyxy[3]
    x_min, y_min, x_max, y_max = int(x_min), int(y_min), int(x_max), int(y_max)
    cropped_img = image_rgb[y_min:y_max, x_min:x_max]

    return cropped_img


def container_predict(image_rgb):
    # предсказываем контейнеры
    results_cont = model_container.predict(image_rgb)
    # получаем координаты рамок с контейнерами
    frame, xyxys, confidences, class_ids = model_container.plot_bbboxes(results_cont, image_rgb)

    # Отобразите результаты
    for result in results_cont:
        boxes = result.boxes  # Получите ограничивающие рамки
        for box in boxes:
            x1, y1, x2, y2 = box.xyxy[0]  # Координаты bbox
            conf = box.conf[0]  # Уверенность
            cls = box.cls[0]  # Класс

            # Нарисуйте bbox на изображении
            cv2.rectangle(image_rgb, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 3)  # Красный цвет
            cv2.putText(image_rgb, f'Class: {int(cls)}, Conf: {conf:.2f}', (int(x1), int(y1) + 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

    # st.write("Найденный контейнер для:", im_name)
    st.image(image_rgb, use_column_width=True)

    return xyxys


def damage_predict(image, xyxys):
    results_dmg = model_container_damage.predict(frame=image, box=xyxys)

    image_rgb = np.array(image)

    cropped_img = get_cropped_img_by_xyxys(image_rgb, xyxys)
    paint_results_dmg_number(cropped_img, results_dmg)

    return results_dmg

    # for result_list in results_dmg:
    #     for result in result_list:
    #         names = result.names
    #         for box in boxes:
    #             class_name = names[int(box.cls[0])]
    #             print('class_name', class_name)
    #             cont_result.append(dmg_translate[class_name])
    # print('cont_result', cont_result)


def find_cont_number(image, xyxys):
    results_number = model_container_number.predict(frame=image, box=xyxys)

    image_rgb = np.array(image)

    cropped_img = get_cropped_img_by_xyxys(image_rgb, xyxys)
    paint_results_dmg_number(cropped_img, results_number)

    return results_number

    # print('cont_result', cont_result)


def find_cont_number_with_manual_wall(idx, image, container_xyxys):
    result_number_1 = None
    if idx == 1 and (st.session_state.wall_type_1_img == "Передняя стенка" or
                     st.session_state.wall_type_1_img == "Задняя стенка"):
        result_number_1 = find_cont_number(image, container_xyxys)
    elif idx == 2 and (st.session_state.wall_type_2_img == "Передняя стенка" or
                       st.session_state.wall_type_2_img == "Задняя стенка"):
        result_number_1 = find_cont_number(image, container_xyxys)
    elif idx == 3 and (st.session_state.wall_type_3_img == "Передняя стенка" or
                       st.session_state.wall_type_3_img == "Задняя стенка"):
        result_number_1 = find_cont_number(image, container_xyxys)

    return result_number_1


def recogn_number(cont_number_detections):
    result_number = 'TCLO 531461 4'  # Пока заглушка

    return result_number


def clean_temp_folder_text_input(temp_images_dir):
    """Удаляет все файлы в указанной папке, очищает поле ввода номера контейнера."""

    for filename in os.listdir(temp_images_dir):
        file_path = os.path.join(temp_images_dir, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Ошибка при удалении {file_path}: {e}")

    # Очищаем поле ввода (через session_state)
    st.session_state.user_number_input = ""


def make_report(results_dmg, result_number):
    cont_result_all = ''
    cont_result = [None, None]

    for list_result_list_number, list_result_list in enumerate(results_dmg):
        side_str = ''
        for result_list in list_result_list:
            side_str = get_wall_str(list_result_list_number, side_str)
            for result in result_list:
                names = result.names
                boxes = result.boxes
                for box in boxes:
                    class_name = names[int(box.cls[0])]
                    side_str += dmg_translate[class_name] + ', '

            cont_result_all += side_str

    # Полный путь к папке reports (относительно текущего скрипта)
    reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
    os.makedirs(reports_dir, exist_ok=True)  # Создаём, если нет

    # Загружаем шаблон
    AOF_template_path = "docs_templates/AOF_template.xlsx"
    wb = load_workbook(AOF_template_path)
    sheet = wb.active   # Получаем активный лист (предполагаем, что это TDSheet)

    # Записываем данные в указанные ячейки
    # R10C2 строка 10, столбец 2 (B10 в Excel)
    sheet.cell(row=10, column=2, value=result_number)
    # R13C1 строка 13, столбец 1 (A13 в Excel)
    sheet.cell(row=13, column=1, value=cont_result_all)

    filename = os.path.join(reports_dir, 'АОФ ' + result_number + '.xlsx')

    # Сохраняем как новый файл
    wb.save(filename)


# Заголовок приложения
st.title("Определение повреждений контейнера")

# Инструкция для пользователя
st.write("Загрузите изображения, чтобы найти повреждения")

# Загрузка изображения
uploaded_files = st.file_uploader("Загрузите изображения и в панели слева выберите ту сторону контейнера,"
                                  " которая в данный момент видна на каждом изображении", type=["jpg", "jpeg", "png"],
                                  accept_multiple_files=True, key="files_uploader")

# names = model.names
if uploaded_files is not None:
    # print('uploaded_files', uploaded_files)
    images = []
    files_list = []
    temp_images_dir = os.path.join(os.path.dirname(__file__), 'temp_images')

    for uploaded_file in uploaded_files:
        # Отображение загруженного изображения
        image = PILImage.open(uploaded_file)
        im_name = uploaded_file.name
        im_filename = 'temp_images\\'+im_name
        # im_filename = os.path.join(temp_images_dir, 'temp_'+im_name+'.jpg')
        # print('im_name_1', im_name)
        # print('im_filename', im_filename)
        with io.BytesIO() as output:
            image.save(output, format="JPEG")  # Adjust format if needed
            contents = output.getvalue()
            with open(im_filename, 'wb') as f:
                f.write(contents)

        files_list.append(im_filename)
        images.append(image)
        st.image(image, use_column_width=True)
        # st.image(image, caption="Загруженное изображение", use_column_width=True)

    # print('files_list', files_list)

    # Кнопка для запуска классификации
    if st.button("Найти повреждения"):
        start = time.time()

        cont_dmg_result = []
        # print('images', images)
        for idx, image in enumerate(images, start=1):
            im_name = uploaded_files[idx-1]
            curr_im_filename = uploaded_files.index(im_name)
            # print('curr_im_filename', curr_im_filename)
            im_name = im_name.name
            # print('im_name', im_name)

            image_bbs = cv2.imread(files_list[curr_im_filename])
            image_rgb = cv2.cvtColor(image_bbs, cv2.COLOR_BGR2RGB)

            # Предсказание контейнера
            container_xyxys = container_predict(image_rgb)
            # Предсказание повреждений
            result_dmg = damage_predict(image, container_xyxys)
            # Предсказание маркировки
            cont_number_detections = find_cont_number_with_manual_wall(idx, image, container_xyxys)
            # Распознавание номера
            result_number = recogn_number(cont_number_detections)

            # Сохраняем данные в session_state для использования вне этого блока
            cont_dmg_result.append(result_dmg)
            st.session_state.result_number = result_number

        st.session_state.result_dmg = cont_dmg_result

        end = time.time() - start
        print('time', np.round(end, 2))


    # Инициализация session_state (если еще нет)
    if "user_number_input" not in st.session_state:
        st.session_state.user_number_input = ""

    # Создаем поле для ввода текста
    user_number_input = st.text_input("Если номер распознан неверно, введите правильный номер контейнера:",
                                      value=st.session_state.user_number_input,  # Значение берется из session_state
                                      key="container_number_input")  # Уникальный ключ для управления полем
    st.session_state.user_number_input = user_number_input

    if st.button("Сформировать отчет"):
        if user_number_input:
            res_dmg, res_number = st.session_state.result_dmg, st.session_state.user_number_input
        else:
            res_dmg, res_number = st.session_state.result_dmg, st.session_state.result_number

        make_report(res_dmg, res_number)
        # Очищаем папку temp_images
        clean_temp_folder_text_input(temp_images_dir)


with st.sidebar:
    st.header("Для каждого из трех изображений выберите сторону контейнера. "
              "Изображения отображены в порядке сверху вниз от первого к последнему.")
    option_1_img = st.radio(
        "На первом изображении",
        ["Передняя стенка", "Задняя стенка", "Левая стенка", "Правая стенка"],
        key="wall_type_1_img"
    )
    option_2_img = st.radio(
        "На втором изображении",
        ["Передняя стенка", "Задняя стенка", "Левая стенка", "Правая стенка"],
        key="wall_type_2_img"
    )
    option_3_img = st.radio(
        "На третьем изображении",
        ["Передняя стенка", "Задняя стенка", "Левая стенка", "Правая стенка"],
        key="wall_type_3_img"
    )
